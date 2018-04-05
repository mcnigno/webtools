from app import db
from flask_appbuilder.models.sqla.interface import SQLAInterface
from flask import flash, send_file, make_response, redirect, url_for

from .models import (Matrix, Document, Unit, Materialclass, Doctype, Partner,
                     Cdrlitem, Documentclass, Mr, Vendor, DocRequests)
#from .views import send_csv
import csv, xlsxwriter
from werkzeug.utils import secure_filename
import uuid
import openpyxl, os



def adddoc3(self, item):
    # Set the Request type
    if item.vendor and item.mr:
        item.request_type = 'vendor'
    
    session = db.session
    matrix = session.query(Matrix)
    if item.unit.unit == '000':
        item_matrix = str.join('-', (item.unit.unit,
                                     item.materialclass.materialclass,
                                     item.doctype.doctype,
                                     # item.sheet,
                                     item.partner.partner))
    else:
        item_matrix = str.join('-', (item.unit.unit,
                                     item.materialclass.materialclass,
                                     item.doctype.doctype,
                                     # item.sheet,
                                     ))

    item_serial = str.join('-', (item.unit.unit,
                                 item.materialclass.materialclass,
                                 item.doctype.doctype,
                                 # item.sheet,
                                 ))

    print('controllo item_ matrix', item_matrix)
    found = False
    for row in matrix:
        print('loop controllo matrix uguali',
              row.matrix, item_matrix,
              item.matrix)
        if row.matrix == item_matrix:
            print('trovate matrix uguali')
            print('row counter prima:', row.counter)
            row.counter += 1
            print('row counter dopo:', row.counter)
            datamodel = SQLAInterface(Matrix, session=session)
            datamodel.edit(row)

            item.matrix_id = row.id
            code = item_serial + "-" + str(row.counter).zfill(5) + "-" + item.sheet

            datamodel = SQLAInterface(Document, session=session)
            doc = Document(docrequests_id=item.id, code=code)
            datamodel.add(doc)

            message = 'Your code is ' + code
            flash(message, category='info')
            found = True
    
    # Matrix Not Found
    if found is False:
        print('Matrix NOT FOUND')

        if str(item.unit) == '000':
            print('found unit 000')
            jv = {
                    'TTSJV': 50000,
                    'TPIT': 60000
                }

            # Add New Matrix
            datamodel = SQLAInterface(Matrix, session=session)

            print('counter', jv['TTSJV'], item.partner, jv[str(item.partner)])
            print('item matrix:', item_matrix)

            matrix = Matrix(counter=jv[str(item.partner)] + 1,
                            matrix=item_matrix)

            print('-----2------')

            datamodel.add(matrix)

            # Add New Doc with quantity jv + 1
            datamodel = SQLAInterface(Document, session=session)
            code = item_serial + "-" + str(jv[str(item.partner)] + 1).zfill(5) + "-" + item.sheet
            doc = Document(docrequests_id=item.id, code=code)

            datamodel.add(doc)
            message = 'Your code is ' + code
            flash(message, category='info')

        else:
            # Add New Matrix
            datamodel = SQLAInterface(Matrix, session=session)
            print('item matrix:', item_matrix)

            print('item matrix after:', item_matrix)
            matrix = Matrix(matrix=item_matrix)
            datamodel.add(matrix)

            # Add New Doc with quantity 1
            datamodel = SQLAInterface(Document, session=session)
            code = item_serial + "-" + "1".zfill(5) + "-" + item.sheet
            doc = Document(docrequests_id=item.id, code=code)

            datamodel.add(doc)
            message = 'Your code is ' + code
            flash(message, category='info')

    db.session.flush()

def bapco(self, item):
    # Set the DB session
    session = db.session
    
    # Set the Request type
    if item.vendor and item.mr:
        item.request_type = 'vendor'
    else:
        item.request_type = 'engineering'
    
    # Set item_matrix based on unit type
    result = db.session.query(Unit).filter(Unit.unit == str(item.unit)).first()
    
    if str(result.unit_type) == 'common':
        print('Match unit type common Found')

        # Add the partner id to the matrix
        item_matrix = str.join('-', (str(item.unit),
                                     str(item.materialclass),
                                     str(item.doctype),
                                     # item.sheet,
                                     str(item.partner)
                                     ))
    else:
        item_matrix = str.join('-', (str(item.unit),
                                     str(item.materialclass),
                                     str(item.doctype),
                                     # item.sheet,
                                     ))
    # Set the bapco base code
    item_serial = str.join('-', (str(item.unit),
                                 str(item.materialclass),
                                 str(item.doctype),
                                 # item.sheet,
                                 ))
    
    # Increment the Matrix counter or create a new one
    print('Matrix to search:', 
    (str(item.unit),
                                     str(item.materialclass),
                                     str(item.doctype),
                                     # item.sheet,
                                     str(item.partner)
                                     )
    )
    matrix = db.session.query(Matrix).filter(Matrix.matrix == item_matrix).first()
    if matrix:
        matrix.counter += 1
        datamodel = SQLAInterface(Matrix, session=session)
        datamodel.edit(matrix)

        item.matrix_id = matrix.id
        code = item_serial + "-" + str(matrix.counter).zfill(5) + "-" + item.sheet

        datamodel = SQLAInterface(Document, session=session)
        doc = Document(docrequests_id=item.id, code=code)
        datamodel.add(doc)

        message = 'Your code is ' + code
        flash(message, category='info')
    else:
        # Create a New Matrix for common units
        if result.unit_type == 'common':

            print('item partner to find: ', item.partner)
            
            partner = db.session.query(Partner).filter(Partner.partner == str(item.partner)).first()
           
            matrix = Matrix(counter=partner.common_start + 1, matrix=str(item_matrix))
            datamodel = SQLAInterface(Matrix, session=session)
            datamodel.add(matrix)
        

            # Add new Doc with quantity partner common start + 1
            datamodel = SQLAInterface(Document, session=session)
            
            code = item_serial + "-" + str(partner.common_start + 1).zfill(5) + "-" + item.sheet
            
            doc = Document(docrequests_id=item.id, code=code)
            datamodel.add(doc)
            message = 'Your code is ' + code
            flash(message, category='info')
        else:
            # Create a new Matrix for standard units
            datamodel = SQLAInterface(Matrix, session=session)
            matrix = Matrix(matrix=item_matrix)
            datamodel.add(matrix)

            # Add new Doc with quantity 1
            datamodel = SQLAInterface(Document, session=session)
            code = item_serial + "-" + "1".zfill(5) + "-" + item.sheet
            doc = Document(docrequests_id=item.id, code=code)

            datamodel.add(doc)
            message = 'Your code is ' + code
            flash(message, category='info')

    db.session.flush()
    return code


def tocsv(self, item,  codes_list):
    print('tocsv FUNCTION')
    print(codes_list)
    filename = 'app/static/csv/bapco_request_'+ str(item.id) + '.csv'
    with open(filename, 'w') as csv_file:
        # writer = csv.writer(csv_file, quoting=csv.QUOTE_ALL)
        writer = csv.writer(csv_file, dialect='excel')
        head = [['Id Code', 'Bapco Code', 'Contractor Code']] 
        writer.writerows(head)
        writer.writerows(codes_list)
        print('inside the with file for CV rendering')
        #print(csv_file)
    
    print('BEFORE file SEND')
    #send_file(filename, as_attachment=True)
    print('file SEND')
    '''
    response = make_response(filename)
    cd = 'attachment; filename=mycsv.csv'
    response.headers['Content-Disposition'] = cd
    response.mimetype = 'text/csv'
    return response
    '''
def toxlsx(self, item,  codes_list):
    print('toXLSX FUNCTION')
    print('this is the code list', codes_list)
    #filename = 'app/static/csv/bapco_request_'+ str(item.id) + '.xlsx'

    workbook = xlsxwriter.Workbook('app/static/csv/bapco_request_'+ str(item.id) + '.xlsx')
    workbook.set_properties({
                            'title':    'Bapco Request spreadsheet',
                            'subject':  'With document properties',
                            'author':   'Quasar DCC Team',
                            'manager':  'Danilo Pacifico',
                            'company':  'Quasar',
                            'category': 'Bapco spreadsheets',
                            'keywords': 'Bapco, Bapco Codes, Properties',
                            'comments': 'Created by webtools.quasarPM.com'})

    worksheet = workbook.add_worksheet(name='Bapco Request List')

    bold = workbook.add_format({'bold': True})
    worksheet.set_column('A:A', 25)
    worksheet.set_column('B:B', 25)
    worksheet.write('A1', 'Bapco Code', bold)
    worksheet.write('B1', 'Contractor Code', bold)
    # start after the header
    row = 1
    col = 0
    print('this is the touple')
    print(tuple(codes_list))
    #worksheet.write(0,0,'something')
    t_list = tuple(codes_list)
    for code in (t_list):
        #print('Looping colist', code, row)
        worksheet.write(row, col, str(code[0]))
        row += 1
    workbook.close()

def codes_to_xlsx(codes_list):
    print('CODES toXLSX FUNCTION')
    print(codes_list)
    #filename = 'app/static/csv/bapco_request_'+ str(item.id) + '.xlsx'
    filename = 'Quasar|' + str(uuid.uuid4()) + '|bapco.xlsx'
    workbook = xlsxwriter.Workbook('app/static/csv/' + filename)
    workbook.set_properties({
                            'title':    'Bapco Request spreadsheet',
                            'subject':  'With document properties',
                            'author':   'Quasar DCC Team',
                            'manager':  'Danilo Pacifico',
                            'company':  'Quasar',
                            'category': 'Bapco spreadsheets',
                            'keywords': 'Bapco, Bapco Codes, Properties',
                            'comments': 'Created by webtools.quasarPM.com'})

    worksheet = workbook.add_worksheet(name='Bapco Request List')

    bold = workbook.add_format({'bold': True})
    worksheet.set_column('A:A', 25)
    worksheet.set_column('B:B', 25)
    worksheet.write('A1', 'Bapco Code', bold)
    worksheet.write('B1', 'Contractor Code', bold)
    # start after the header
    row = 1
    col = 0
    print(tuple(codes_list))
    #worksheet.write(0,0,'something')
    t_list = tuple(codes_list)
    for code in (t_list):
        print('Looping colist', code, row)
        worksheet.write(row, col, str(code[0]))
        row += 1
    workbook.close()
    return filename

def update_from_xlsx(file):
    session = db.session
    print('update FUNCTION!')
    book = openpyxl.load_workbook(file)
    sheet = book.active
    a1 = sheet['A1']
    print(a1.value)
    reserved_list = []
    updated_list = []
    for row in sheet.iter_rows(min_row=2):
        bapco_code = row[0].value
        oldcode = row[1].value
        print(bapco_code, oldcode)
        doc = db.session.query(Document).filter(Document.code == str(bapco_code)).first()

        if doc and doc.oldcode == 'empty':
            print('this is the ID' ,doc.id)
            datamodel = SQLAInterface(Document, session=session)
            print('BEFORE oldcode', doc.oldcode)
            doc.oldcode = oldcode
            updated_list.append([doc.id, doc.code, doc.oldcode])
            datamodel.edit(doc)

        else:
            reserved_list.append([doc.id, doc.code, doc.oldcode])
    return reserved_list, updated_list


def setting_update(file):
    session = db.session
    print('Setting Update start')
    book = openpyxl.load_workbook(file)
    sheet = book.active
    set_class = sheet['A1'].value
    print('the setting class:', set_class)

    reserved_list = []
    updated_list = []

    for row in sheet.iter_rows(min_row=2):
        param = row[0].value
        name = row[1].value
        desc = row[2].value
        print(param, name, desc)
        set_dict = {
            'Unit': [Unit, Unit.unit, 'unit'],
            'Materialclass': [Materialclass, Materialclass.materialclass],
            'Doctype': [Doctype, Doctype.doctype],
            'Cdrlitem': [Cdrlitem, Cdrlitem.cdrlitem],
            'Documentclass': [Documentclass, Documentclass.documentclass],
            'Mr': [Mr, Mr.mr],
            'Vendor': [Vendor, Vendor.vendor],
            'Partner': [Partner, Partner.partner]
        }
        # Update the setting if a param already exist
        tmp_class = set_dict[set_class][0]
        tmp_param = set_dict[set_class][1]
        my_class = db.session.query(tmp_class).filter(tmp_param == str(param)).first()
        
        datamodel = SQLAInterface(tmp_class, session=session)
        
        if my_class:
            print(my_class)
            my_class.name = name
            my_class.description = desc
            datamodel.edit(my_class)
            updated_list.append([my_class.id, my_class.name, my_class.description])
        else:
            # or Create new record in setting
            
            my_class = tmp_class()
            if set_class == 'Unit':
                my_class.unit = param
            elif set_class == 'Materialclass':
                my_class.materialclass = param
            elif set_class == 'Doctype':
                my_class.doctype = param
            elif set_class == 'Cdrlitem':
                my_class.cdrlitem = param
            elif set_class == 'Documentclass':
                my_class.documentclass = param
            elif set_class == 'Mr':
                my_class.mr = param
            elif set_class == 'Vendor':
                my_class.vendor = param
            elif set_class == 'Partner':
                my_class.partner = param
            
            else:
                return print('Setting Class NOT Found')
            
            my_class.name = name
            my_class.description = desc
            datamodel.add(my_class)
            reserved_list.append([my_class.id, my_class.name, my_class.description])

    return reserved_list, updated_list

def old_codes_update(self, file):
    book = openpyxl.load_workbook(file)
    sheet = book.active
    #header = sheet['A1':'M1']
    '''
    unit = header[0][0].value
    materialclass = header[0][1].value
    doctype = header[0][2].value
    bapco_code = header[0][5].value
    description = header[0][6].value
    oldcode = header[0][7].value
    cdrlitem = header[0][8].value
    documentclass =header[0][9].value
    note = header[0][10].value
    trasmittal = header[0][11].value
    '''
    session = db.session
    datamodel = SQLAInterface(DocRequests, session=session)

    for row in sheet.iter_rows(min_row=2):
        req = DocRequests()
        req.quantity = 1
        req.unit = Unit(unit=row[0].value) 
        req.materialclass = Materialclass(materialclass=row[1].value) 
        req.doctype = Doctype(doctype=row[2].value) 
        req.cdrlitem = Cdrlitem(cdrlitem=row[8].value) 
        req.documentclass = Documentclass(documentclass=row[9].value) 
        req.partner = Partner(partner=row[12].value) 
        #req.oldcode = row[7].value
        print('search for Unit:', req.unit.unit)
        #code = bapco(self, req)
        #row[14] = code
        datamodel.add(req)
        
    reserved_list = ['coming soon...']
    updated_list = ['coming soon...']
    book.close()
    return reserved_list, updated_list

#xls = open('bapco_codes.xlsx','rb')
def old_codes(self, file):
    book = openpyxl.load_workbook(file)
    sheet = book.active
    session = db.session
    # Create the datamodel
    datamodel = SQLAInterface(DocRequests, session=session)
    first_req = DocRequests(unit=Unit(unit='QSR'), materialclass=Materialclass(materialclass='1'), doctype=Doctype(doctype='QSR'), partner=Partner(partner='QSR'))
    datamodel.add(first_req)
    

    
    found_list = []
    not_found_list = []
    for row in sheet.iter_rows(min_row=2):
        check = True
        datamodel = SQLAInterface(DocRequests, session=session)
        req = DocRequests()
        #req.id = 1
       
        # Unit Query
        if row[0].value:
            try:
                unit_id = session.query(Unit).filter(Unit.unit == row[0].value).first()
                req.unit_id = unit_id.id
                req.unit = unit_id.unit
                #req.unit.unit = unit_id.unit
                
                if unit_id.unit_type == '000':
                    req.unit_type = 'common'

                found_list.append(['Unit', req.unit_id, row[0].value])
            except:
                check = False
                not_found_list.append(['bapco: '+row[5].value+' for '+row[7].value,'Unit Not Found',row[0].value ])

        # Materialclass Query
        if row[1].value:
            try:
                mat_id = session.query(Materialclass).filter(Materialclass.materialclass == row[1].value).first()
                req.materialclass_id = mat_id.id
                req.materialclass = mat_id.materialclass
                found_list.append(['Materialclass', req.materialclass_id, row[1].value])
            except:
                check = False
                not_found_list.append(['bapco: '+row[5].value+' for '+row[7].value,'Material class Not Found',row[1].value ])
        
        # Doctype Query
        if row[2].value:
            try:
                doc_id = session.query(Doctype).filter(Doctype.doctype == row[2].value).first()
                req.doctype_id = doc_id.id
                req.doctype = doc_id.doctype
                found_list.append(['Doctype', req.doctype_id, row[2].value])
            except:
                check = False
                not_found_list.append(['bapco: '+row[5].value+' for '+row[7].value,'Doctype Not Found',row[2].value ])

        # CdrlItem Query
        if row[8].value:
            try:
                cdrl_id = session.query(Cdrlitem).filter(Cdrlitem.cdrlitem == row[8].value).first()
                req.cdrlitem_id = cdrl_id.id
                req.cdrlitem = cdrl_id.cdrlitem
                found_list.append(['Cdrlitem', req.cdrlitem_id, row[8].value])
            except:
                #check = False
                not_found_list.append(['bapco: '+row[5].value+' for '+row[7].value,'Cdrl Not Found',row[8].value ])
        
        # Documentclass Query
        if row[9].value:
            try:
                dc_id = session.query(Documentclass).filter(Documentclass.documentclass == row[9].value).first()
                req.documentclass_id = dc_id.id
                req.documentclass = dc_id.documentclass
                found_list.append(['Documentclass', req.documentclass_id, row[9].value])
            except:
                #check = False
                not_found_list.append(['bapco: '+row[5].value+' for '+row[7].value,'Document class Not Found',row[9].value ])

        # Partner Query
        if row[12].value:
            try:
                
                pa_id = session.query(Partner).filter(Partner.partner == row[12].value).first()
                req.partner_id = pa_id.id
                req.partner = pa_id.partner
                found_list.append(['Partner', req.partner_id, row[12].value])
            except:
                check = False
                not_found_list.append(['bapco: '+row[5].value+' for '+row[7].value,'Partner Not Found', row[12].value ])
        
        # Add Sheet 001
        req.sheet = '001'
        #
        # 
        # ADD the first Request with id = 1
        #
        '''
        datamodel = SQLAInterface(DocRequests, session=session)
        first_req = DocRequests(unit='001', materialclass='A', doctype='HDD', partner='QSR')
        datamodel.add(first_req)
        req.id = 1
        '''
        
        if check is True:
            code = bapco(self, req)
            ask_bapco_code = row[13].value
            row[13].value = code
            
            
            print('Added DocRequstest:', req.unit_id, req.materialclass_id, req.doctype_id,
                    req.cdrlitem_id, req.documentclass_id, req.partner_id)
            print('Ask Bapco:', code)
            print(' Your Bapco Code: ',row[5].value)
            print('Your Code: ',row[7].value )

            
        else:
            print('Not Added DocRequstest:', req.unit_id, req.materialclass_id, req.doctype_id,
                    req.cdrlitem_id, req.documentclass_id, req.partner_id)
            print('Wrong Request: ',row[0].value,row[1].value,row[2].value,row[8].value,row[9].value,row[12].value )
    
    result_file = 'app/static/csv/upload_results.xlsx'
    book.save(result_file)


    '''
    print('Found List')
    for i in found_list:
        print(i[0],i[1], i[2])
    
    print('NOT Found List')
    for i in not_found_list:
        print(i[0],i[1], i[2])
    '''
    return not_found_list, found_list, result_file 